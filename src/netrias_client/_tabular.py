"""Represent tabular source files without losing column identity.

CSV, TSV, and XLSX are file formats at the SDK boundary. Inside the client,
columns are identified by position-derived keys so duplicate display headers
remain distinct.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import cast

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as OpenPyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet


class TabularFormat(str, Enum):
    CSV = "csv"
    TSV = "tsv"
    XLSX = "xlsx"

    @property
    def delimiter(self) -> str:
        if self == TabularFormat.CSV:
            return ","
        if self == TabularFormat.TSV:
            return "\t"
        raise ValueError("XLSX does not have a text delimiter")

    @property
    def suffix(self) -> str:
        return f".{self.value}"

    @property
    def content_type(self) -> str:
        if self == TabularFormat.CSV:
            return "text/csv"
        if self == TabularFormat.TSV:
            return "text/tab-separated-values"
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class WorkbookSheet:
    name: str
    index: int


@dataclass(frozen=True)
class TabularColumn:
    key: str
    index: int
    header: str


@dataclass(frozen=True)
class TabularDataset:
    columns: list[TabularColumn]
    rows: list[list[str]]
    source_format: TabularFormat
    sheet_name: str | None = None

    @property
    def headers(self) -> list[str]:
        return [column.header for column in self.columns]

    def backend_column_names(self) -> list[str]:
        return [_backend_column_name(column) for column in self.columns]


SUPPORTED_TABULAR_SUFFIXES: dict[str, TabularFormat] = {
    ".csv": TabularFormat.CSV,
    ".tsv": TabularFormat.TSV,
    ".xlsx": TabularFormat.XLSX,
}
SUPPORTED_TABULAR_FORMATS: tuple[TabularFormat, ...] = tuple(SUPPORTED_TABULAR_SUFFIXES.values())
_CSV_CONTENT_TYPES = {"text/csv", "application/csv", "application/vnd.ms-excel"}
_TSV_CONTENT_TYPES = {"text/tab-separated-values", "text/tsv", "text/plain"}
_XLSX_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
}
_CONTENT_TYPE_FORMATS: dict[str, TabularFormat] = {
    **dict.fromkeys(_CSV_CONTENT_TYPES, TabularFormat.CSV),
    **dict.fromkeys(_TSV_CONTENT_TYPES, TabularFormat.TSV),
    **dict.fromkeys(_XLSX_CONTENT_TYPES, TabularFormat.XLSX),
}


def column_key_for_index(index: int) -> str:
    return f"col_{index:04d}"


def column_index_from_key(column_key: str) -> int | None:
    prefix = "col_"
    if not column_key.startswith(prefix):
        return None
    suffix = column_key[len(prefix):]
    if not suffix.isdigit():
        return None
    return int(suffix)


def tabular_format_for_path(path: Path) -> TabularFormat:
    file_format = SUPPORTED_TABULAR_SUFFIXES.get(path.suffix.lower())
    if file_format is None:
        raise ValueError(f"unsupported tabular file extension: {path.suffix}")
    return file_format


def get_tabular_format(path: Path, content_type: str | None = None) -> TabularFormat:
    file_format = SUPPORTED_TABULAR_SUFFIXES.get(path.suffix.lower())
    if file_format is not None:
        return file_format

    normalized_content_type = (content_type or "").lower()
    content_type_format = _CONTENT_TYPE_FORMATS.get(normalized_content_type)
    if content_type_format is not None:
        return content_type_format

    raise ValueError(f"unsupported tabular file extension: {path.suffix or '<none>'}")


def is_supported_tabular_content_type(content_type: str, file_format: TabularFormat) -> bool:
    normalized = content_type.lower()
    if file_format == TabularFormat.CSV:
        return normalized in _CSV_CONTENT_TYPES
    if file_format == TabularFormat.TSV:
        return normalized in _TSV_CONTENT_TYPES
    return normalized in _XLSX_CONTENT_TYPES


def list_workbook_sheets(path: Path) -> list[WorkbookSheet]:
    if tabular_format_for_path(path) != TabularFormat.XLSX:
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [WorkbookSheet(name=name, index=index) for index, name in enumerate(workbook.sheetnames)]
    finally:
        workbook.close()


def read_tabular(path: Path, sheet_name: str | None = None) -> TabularDataset:
    source_format = tabular_format_for_path(path)
    if source_format == TabularFormat.XLSX:
        return _read_xlsx(path, sheet_name)
    return _read_delimited_text(path, source_format)


def write_tabular(path: Path, dataset: TabularDataset, template_path: Path | None = None) -> None:
    if dataset.source_format == TabularFormat.XLSX:
        _write_xlsx(path, dataset, template_path)
        return
    _write_delimited_text(path, dataset)


def dataset_from_rows(
    *,
    columns: list[TabularColumn] | None = None,
    headers: list[str] | None = None,
    rows: list[list[str]],
    source_format: TabularFormat,
    sheet_name: str | None = None,
) -> TabularDataset:
    base_headers = [column.header for column in columns] if columns is not None else (headers or [])
    width = max([len(base_headers), *(len(row) for row in rows)] or [0])
    normalized_headers = [*base_headers, *([""] * (width - len(base_headers)))]
    output_columns = _columns_for_headers(normalized_headers, columns)
    return TabularDataset(
        columns=output_columns,
        rows=[_normalize_row(row, width) for row in rows],
        source_format=source_format,
        sheet_name=sheet_name,
    )


def csv_bytes_to_dataset(
    content: bytes,
    source_format: TabularFormat,
    sheet_name: str | None = None,
) -> TabularDataset:
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    return dataset_from_rows(headers=headers, rows=data_rows, source_format=source_format, sheet_name=sheet_name)


def _read_delimited_text(path: Path, source_format: TabularFormat) -> TabularDataset:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=source_format.delimiter)
        default_headers: list[str] = []
        headers = next(reader, default_headers)
        raw_rows = [list(row) for row in reader]

    return dataset_from_rows(headers=headers, rows=raw_rows, source_format=source_format)


def _write_delimited_text(path: Path, dataset: TabularDataset) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter=dataset.source_format.delimiter, lineterminator="\n")
        writer.writerow(dataset.headers)
        writer.writerows(dataset.rows)


def _read_xlsx(path: Path, sheet_name: str | None) -> TabularDataset:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _select_sheet(workbook, sheet_name)
        rows = [[_cell_to_string(value) for value in row] for row in sheet.iter_rows(values_only=True)]
        selected_sheet_name = sheet.title
    finally:
        workbook.close()

    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    return dataset_from_rows(
        headers=headers,
        rows=data_rows,
        source_format=TabularFormat.XLSX,
        sheet_name=selected_sheet_name,
    )


def _write_xlsx(path: Path, dataset: TabularDataset, template_path: Path | None) -> None:
    workbook = load_workbook(template_path) if template_path else Workbook()
    sheet = _writable_sheet(workbook, dataset.sheet_name)
    _replace_sheet_values(sheet, dataset)
    workbook.save(path)


def _select_sheet(workbook: OpenPyxlWorkbook, sheet_name: str | None) -> Worksheet:
    if sheet_name is None:
        return cast(Worksheet, workbook[workbook.sheetnames[0]])
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"unknown worksheet: {sheet_name}; available worksheets: {available}")
    return cast(Worksheet, workbook[sheet_name])


def _writable_sheet(workbook: OpenPyxlWorkbook, sheet_name: str | None) -> Worksheet:
    active = workbook.active
    candidate_name = sheet_name or (active.title if active is not None else "Sheet1")
    if candidate_name in workbook.sheetnames:
        return cast(Worksheet, workbook[candidate_name])
    sheet = cast(Worksheet, workbook.create_sheet(candidate_name))
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1 and workbook["Sheet"].max_row == 1:
        workbook.remove(workbook["Sheet"])
    return sheet


def _replace_sheet_values(sheet: Worksheet, dataset: TabularDataset) -> None:
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)
    sheet.append(dataset.headers)
    for row in dataset.rows:
        sheet.append(row)


def _cell_to_string(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _backend_column_name(column: TabularColumn) -> str:
    return f"{column.key}__{_safe_header_token(column.header)}"


def _columns_for_headers(headers: list[str], columns: list[TabularColumn] | None) -> list[TabularColumn]:
    if columns is None:
        return [
            TabularColumn(key=column_key_for_index(index), index=index, header=header)
            for index, header in enumerate(headers)
        ]
    if len(columns) >= len(headers):
        return columns[:len(headers)]
    extra_columns = [
        TabularColumn(key=column_key_for_index(index), index=index, header=headers[index])
        for index in range(len(columns), len(headers))
    ]
    return [*columns, *extra_columns]


def _safe_header_token(header: str) -> str:
    cleaned = "".join(char if char.isalnum() else "_" for char in header.strip().lower())
    collapsed = "_".join(part for part in cleaned.split("_") if part)
    return collapsed or "blank"


def _normalize_row(row: list[str], width: int) -> list[str]:
    if len(row) >= width:
        return row
    return [*row, *([""] * (width - len(row)))]


__all__ = [
    "SUPPORTED_TABULAR_FORMATS",
    "SUPPORTED_TABULAR_SUFFIXES",
    "TabularColumn",
    "TabularDataset",
    "TabularFormat",
    "WorkbookSheet",
    "column_index_from_key",
    "column_key_for_index",
    "csv_bytes_to_dataset",
    "dataset_from_rows",
    "get_tabular_format",
    "is_supported_tabular_content_type",
    "list_workbook_sheets",
    "read_tabular",
    "tabular_format_for_path",
    "write_tabular",
]
