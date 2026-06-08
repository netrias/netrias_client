"""Exercise harmonization workflows against mocked transports.

'why': validate success, failure, and transport error handling end-to-end
"""
from __future__ import annotations

import gzip
import json
from pathlib import Path
from typing import cast

import httpx
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from netrias_client import ColumnKeyedManifestPayload, NetriasClient
from netrias_client._errors import NetriasAPIUnavailable
from netrias_client._models import HarmonizationResult

from ._utils import EXTERNAL_VERSION_NUMBER, install_mock_transport, job_success, json_failure, transport_error


def _active_sheet(workbook: Workbook) -> Worksheet:
    return cast(Worksheet, workbook.active)


def _create_sheet(workbook: Workbook, title: str) -> Worksheet:
    return cast(Worksheet, workbook.create_sheet(title))


def _workbook_cell_value(path: Path, sheet_name: str, cell: str) -> object:
    workbook = load_workbook(path, data_only=True)
    sheet = cast(Worksheet, workbook[sheet_name])
    return cast(object, sheet[cell].value)


def test_harmonize_streaming_success(
    configured_client: NetriasClient,
    sample_csv_path: Path,
    sample_manifest_path: Path,
    output_directory: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Write harmonized output and return a success result when the API streams CSV bytes."""

    # Given: a harmonization client with a successful streaming API response
    capture = job_success(chunks=(b"col1,col2\n", b"7,8\n"))
    install_mock_transport(monkeypatch, capture)
    expected_output = output_directory / "sample.harmonized.csv"
    assert not expected_output.exists()
    assert capture.requests == []

    # When: the user harmonizes a CSV file
    result: HarmonizationResult = configured_client.harmonize(
        source_path=sample_csv_path,
        manifest=sample_manifest_path,
        data_commons_key="ccdi",
        external_version_number=EXTERNAL_VERSION_NUMBER,
        output_path=output_directory,
    )

    # Then: the client writes the harmonized file and reports the completed job
    assert result.status == "succeeded"
    assert result.file_path == expected_output
    assert result.job_id == "job-123"
    assert expected_output.exists()
    assert expected_output.read_text(encoding="utf-8") == "col1,col2\n7,8\n"
    assert len(capture.requests) == 3
    submit_request = capture.requests[0]
    poll_request = capture.requests[1]
    final_request = capture.requests[2]

    assert submit_request.method == "POST"
    assert submit_request.url.path.endswith("/v1/jobs/harmonize")
    assert submit_request.headers.get("x-api-key") == "test-api-key"
    submit_body = _decode_submit_body(submit_request)
    assert submit_body.get("data_commons_key") == "ccdi"
    assert submit_body.get("use_cache") is True
    assert submit_body.get("external_version_number") == EXTERNAL_VERSION_NUMBER
    assert poll_request.method == "GET"
    assert "/v1/jobs/" in poll_request.url.path
    assert final_request.method == "GET"


def test_harmonize_handles_api_failure(
    configured_client: NetriasClient,
    sample_csv_path: Path,
    sample_manifest_path: Path,
    output_directory: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Return a failed HarmonizationResult when the API responds with an error payload."""

    # Given: a harmonization API response that rejects the submitted job
    capture = json_failure({"message": "invalid mapping"}, status_code=400)
    install_mock_transport(monkeypatch, capture)
    expected_output = output_directory / "sample.harmonized.csv"
    assert not expected_output.exists()
    assert capture.requests == []

    # When: the user tries to harmonize a CSV file
    result: HarmonizationResult = configured_client.harmonize(
        source_path=sample_csv_path,
        manifest=sample_manifest_path,
        data_commons_key="ccdi",
        external_version_number=EXTERNAL_VERSION_NUMBER,
        output_path=output_directory,
    )

    # Then: the client returns a failed result without writing output
    assert result.status == "failed"
    assert result.description == "invalid mapping"
    assert result.file_path == expected_output
    assert result.job_id is None
    assert not expected_output.exists()
    assert len(capture.requests) == 1


def test_harmonize_raises_on_transport_error(
    configured_client: NetriasClient,
    sample_csv_path: Path,
    sample_manifest_path: Path,
    output_directory: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Raise NetriasAPIUnavailable when the transport layer fails."""

    # Given: a harmonization API transport that cannot be reached
    capture = transport_error(httpx.ConnectError("boom"))
    install_mock_transport(monkeypatch, capture)
    expected_output = output_directory / "sample.harmonized.csv"
    assert not expected_output.exists()
    assert capture.requests == []

    # When: the user tries to harmonize a CSV file
    with pytest.raises(NetriasAPIUnavailable):
        _ = configured_client.harmonize(
            source_path=sample_csv_path,
            manifest=sample_manifest_path,
            data_commons_key="ccdi",
            external_version_number=EXTERNAL_VERSION_NUMBER,
            output_path=output_directory,
        )

    # Then: the client reports the transport error and leaves no output behind
    assert len(capture.requests) == 1
    assert not expected_output.exists()


def test_harmonize_accepts_manifest_mapping(
    configured_client: NetriasClient,
    sample_csv_path: Path,
    sample_manifest_mapping: dict[str, object],
    output_directory: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Allow callers to provide manifest data without writing a file."""

    # Given: a manifest mapping supplied directly by the caller
    capture = job_success(chunks=(b"col1,col2\n", b"7,8\n"))
    install_mock_transport(monkeypatch, capture)
    expected_output = output_directory / "sample.harmonized.csv"
    assert not expected_output.exists()
    assert capture.requests == []

    # When: the user harmonizes with the in-memory manifest mapping
    result: HarmonizationResult = configured_client.harmonize(
        source_path=sample_csv_path,
        manifest=sample_manifest_mapping,
        data_commons_key="ccdi",
        external_version_number=EXTERNAL_VERSION_NUMBER,
        output_path=output_directory,
    )

    # Then: the client writes harmonized output using that mapping
    assert result.status == "succeeded"
    assert result.file_path == expected_output
    assert expected_output.exists()


def test_harmonize_tsv_writes_tsv_output(
    configured_client: NetriasClient,
    output_directory: Path,
    monkeypatch: pytest.MonkeyPatch,
    sample_tsv_path: Path,
) -> None:
    """TSV input produces TSV output while the remote CSV stream is converted."""

    manifest: ColumnKeyedManifestPayload = {
        "column_mappings": {
            "col_0000": {
                "column_name": "name",
                "cde_key": "first_name",
                "cde_id": 10,
                "harmonization": "harmonizable",
                "alternatives": [],
            }
        }
    }
    capture = job_success(chunks=(b'name,note\nAlice,"keeps, comma"\n',))
    install_mock_transport(monkeypatch, capture)

    result = configured_client.harmonize(
        source_path=sample_tsv_path,
        manifest=manifest,
        data_commons_key="ccdi",
        external_version_number=EXTERNAL_VERSION_NUMBER,
        output_path=output_directory,
    )

    expected = output_directory / "sample.harmonized.tsv"
    assert result.file_path == expected
    assert expected.read_text(encoding="utf-8") == "name\tnote\nAlice\tkeeps, comma\n"

    submit_request = capture.requests[0]
    envelope = _decode_submit_body(submit_request)
    document = cast(dict[str, object], envelope["document"])
    assert document["header"] == ["name", "note"]
    assert document["rows"] == [["Alice", "keeps, comma"]]
    column_mappings = cast(list[dict[str, object]], envelope["column_mappings"])
    assert column_mappings[0]["cde_key"] == "first_name"


def test_harmonize_xlsx_writes_xlsx_output_for_selected_sheet(
    configured_client: NetriasClient,
    output_directory: Path,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """XLSX input produces XLSX output while updating only the selected worksheet."""

    source = tmp_path / "patients.xlsx"
    workbook = Workbook()
    keep = _active_sheet(workbook)
    keep.title = "Keep"
    keep.append(["status"])
    keep.append(["unchanged"])
    patients = _create_sheet(workbook, "Patients")
    patients.append(["name", "note"])
    patients.append(["Alice", "old"])
    workbook.save(source)

    manifest: ColumnKeyedManifestPayload = {
        "column_mappings": {
            "col_0000": {
                "column_name": "name",
                "cde_key": "first_name",
                "cde_id": 10,
                "harmonization": "harmonizable",
                "alternatives": [],
            }
        }
    }
    capture = job_success(chunks=(b"name,note\nAlice,updated\n",))
    install_mock_transport(monkeypatch, capture)

    result = configured_client.harmonize(
        source_path=source,
        manifest=manifest,
        data_commons_key="ccdi",
        external_version_number=EXTERNAL_VERSION_NUMBER,
        output_path=output_directory,
        sheet_name="Patients",
    )

    expected = output_directory / "patients.harmonized.xlsx"
    assert result.file_path == expected
    output_workbook = load_workbook(expected, data_only=True)
    assert output_workbook.sheetnames == ["Keep", "Patients"]
    assert _workbook_cell_value(expected, "Keep", "A2") == "unchanged"
    assert _workbook_cell_value(expected, "Patients", "B2") == "updated"

    submit_request = capture.requests[0]
    envelope = _decode_submit_body(submit_request)
    document = cast(dict[str, object], envelope["document"])
    assert document["sheetName"] == "Patients"
    assert document["rows"] == [["Alice", "old"]]


def test_harmonize_writes_manifest_when_requested(
    configured_client: NetriasClient,
    sample_csv_path: Path,
    sample_manifest_mapping: dict[str, object],
    output_directory: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Persist manifest data to disk when a destination path is supplied."""

    # Given: a manifest mapping and a requested manifest output path
    capture = job_success(chunks=(b"col1,col2\n", b"7,8\n"))
    install_mock_transport(monkeypatch, capture)

    manifest_output = tmp_path / "manifest.json"
    assert not manifest_output.exists()
    assert capture.requests == []

    # When: the user harmonizes with manifest output enabled
    _ = configured_client.harmonize(
        source_path=sample_csv_path,
        manifest=sample_manifest_mapping,
        data_commons_key="ccdi",
        external_version_number=EXTERNAL_VERSION_NUMBER,
        output_path=output_directory,
        manifest_output_path=manifest_output,
    )

    # Then: the client writes the input manifest mapping to that path
    assert manifest_output.exists()
    loaded = cast(dict[str, object], json.loads(manifest_output.read_text(encoding="utf-8")))
    assert loaded == sample_manifest_mapping


def test_harmonize_downloads_manifest_from_status_payload(
    configured_client: NetriasClient,
    sample_csv_path: Path,
    sample_manifest_path: Path,
    output_directory: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Download the manifest artifact when the completed job reports one."""

    # Given: a completed job that exposes a manifest artifact URL whose filename should not be trusted
    manifest_url = "https://mock.netrias/sample.csv"
    capture = job_success(
        chunks=(b"col1,col2\n", b"7,8\n"),
        manifest_url=manifest_url,
        manifest_chunks=(b"PAR1",),
    )
    install_mock_transport(monkeypatch, capture)
    first_manifest_path = output_directory / "sample.harmonized.manifest.parquet"
    _ = first_manifest_path.write_bytes(b"existing manifest")
    expected_manifest_path = output_directory / "sample.harmonized.manifest.v1.parquet"
    assert not expected_manifest_path.exists()
    assert capture.requests == []

    # When: the user harmonizes a CSV file
    result = configured_client.harmonize(
        source_path=sample_csv_path,
        manifest=sample_manifest_path,
        data_commons_key="ccdi",
        external_version_number=EXTERNAL_VERSION_NUMBER,
        output_path=output_directory,
    )

    # Then: the client downloads the manifest artifact to a versioned SDK-owned path
    assert result.manifest_path == expected_manifest_path
    assert first_manifest_path.read_bytes() == b"existing manifest"
    assert expected_manifest_path.read_bytes() == b"PAR1"
    assert str(capture.requests[2].url) == manifest_url


def test_harmonize_can_disable_cache(
    configured_client: NetriasClient,
    sample_csv_path: Path,
    sample_manifest_path: Path,
    output_directory: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Send use_cache=false when callers disable cache use."""

    # Given: a harmonization request that has not been submitted yet
    capture = job_success(chunks=(b"col1,col2\n", b"7,8\n"))
    install_mock_transport(monkeypatch, capture)
    assert capture.requests == []

    # When: the user disables cache use for harmonization
    _ = configured_client.harmonize(
        source_path=sample_csv_path,
        manifest=sample_manifest_path,
        data_commons_key="ccdi",
        external_version_number=EXTERNAL_VERSION_NUMBER,
        output_path=output_directory,
        use_cache=False,
    )

    # Then: the submit payload asks the backend not to use cached model results
    submit_body = _decode_submit_body(capture.requests[0])
    assert submit_body.get("use_cache") is False


@pytest.mark.asyncio
async def test_harmonize_async_success(
    configured_client: NetriasClient,
    sample_csv_path: Path,
    sample_manifest_path: Path,
    output_directory: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Async client method returns a successful result."""

    # Given: an async harmonization client with a successful streaming API response
    capture = job_success(chunks=(b"col1,col2\n", b"7,8\n"))
    install_mock_transport(monkeypatch, capture)
    expected_output = output_directory / "sample.harmonized.csv"
    assert not expected_output.exists()
    assert capture.requests == []

    # When: the user harmonizes a CSV file asynchronously
    result = await configured_client.harmonize_async(
        source_path=sample_csv_path,
        manifest=sample_manifest_path,
        data_commons_key="ccdi",
        external_version_number=EXTERNAL_VERSION_NUMBER,
        output_path=output_directory,
    )

    # Then: the async method writes the output and returns the completed job id
    assert result.status == "succeeded"
    assert result.file_path == expected_output
    assert result.job_id == "job-123"
    assert expected_output.exists()


def _decode_submit_body(request: httpx.Request) -> dict[str, object]:
    raw = gzip.decompress(request.content)
    return cast(dict[str, object], json.loads(raw.decode("utf-8")))
