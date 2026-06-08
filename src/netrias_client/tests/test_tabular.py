"""Exercise first-class tabular file handling.

'why': CSV and TSV are edge formats; the SDK should preserve positional column
identity internally so duplicate headers never collapse into dict keys.
"""
from __future__ import annotations

from pathlib import Path
from typing import cast

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from netrias_client import (
    TabularFormat,
    dataset_from_rows,
    list_workbook_sheets,
    read_tabular,
    write_tabular,
)


def _active_sheet(workbook: Workbook) -> Worksheet:
    return cast(Worksheet, workbook.active)


def _create_sheet(workbook: Workbook, title: str) -> Worksheet:
    return cast(Worksheet, workbook.create_sheet(title))


def _workbook_cell_value(path: Path, sheet_name: str, cell: str) -> object:
    workbook = load_workbook(path, data_only=True)
    sheet = cast(Worksheet, workbook[sheet_name])
    return cast(object, sheet[cell].value)


def test_read_tabular_preserves_duplicate_tsv_headers() -> None:
    """A TSV with duplicate headers keeps separate positional column keys."""

    # Given: a TSV fixture with duplicate display headers and comma-containing values
    source = Path(__file__).parent / "fixtures" / "duplicate_headers.tsv"
    assert "," in source.read_text(encoding="utf-8")

    # When: the SDK reads it as tabular data
    dataset = read_tabular(source)

    # Then: identity is positional and data is not split on commas
    assert dataset.source_format == TabularFormat.TSV
    assert [column.key for column in dataset.columns] == ["col_0000", "col_0001", "col_0002"]
    assert dataset.headers == ["name", "name", "note"]
    assert dataset.rows == [["Alice", "Smith", "keeps, comma"]]


def test_read_tabular_selects_xlsx_sheet_and_preserves_duplicate_headers(tmp_path: Path) -> None:
    """XLSX input is read from the selected worksheet without collapsing duplicate headers."""

    # Given: a workbook with two sheets and duplicate headers on the second sheet
    source = tmp_path / "workbook.xlsx"
    workbook = Workbook()
    first = _active_sheet(workbook)
    first.title = "First"
    first.append(["ignored"])
    first.append(["nope"])
    second = _create_sheet(workbook, "Patients")
    second.append(["name", "name", "note"])
    second.append(["Alice", "Smith", "keeps, comma"])
    workbook.save(source)
    assert list_workbook_sheets(source)[0].name == "First"

    # When: the second sheet is selected
    dataset = read_tabular(source, sheet_name="Patients")

    # Then: values come from that sheet and duplicate headers keep distinct keys
    assert dataset.source_format == TabularFormat.XLSX
    assert dataset.sheet_name == "Patients"
    assert dataset.headers == ["name", "name", "note"]
    assert [column.key for column in dataset.columns] == ["col_0000", "col_0001", "col_0002"]
    assert dataset.rows == [["Alice", "Smith", "keeps, comma"]]


def test_write_tabular_xlsx_updates_selected_sheet_and_keeps_other_sheets(tmp_path: Path) -> None:
    """XLSX output updates only the selected sheet when a template workbook is provided."""

    # Given: a template workbook with a sheet that should remain untouched
    template = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    first = _active_sheet(workbook)
    first.title = "Keep"
    first.append(["status"])
    first.append(["unchanged"])
    selected = _create_sheet(workbook, "Patients")
    selected.append(["name", "note"])
    selected.append(["old", "old note"])
    workbook.save(template)
    dataset = dataset_from_rows(
        headers=["name", "note"],
        rows=[["Alice", "updated"]],
        source_format=TabularFormat.XLSX,
        sheet_name="Patients",
    )

    # When: the dataset is written back as XLSX
    write_tabular(output, dataset, template_path=template)

    # Then: the selected sheet is replaced and the other sheet stays present
    output_workbook = load_workbook(output, data_only=True)
    assert output_workbook.sheetnames == ["Keep", "Patients"]
    assert _workbook_cell_value(output, "Keep", "A2") == "unchanged"
    assert _workbook_cell_value(output, "Patients", "A1") == "name"
    assert _workbook_cell_value(output, "Patients", "A2") == "Alice"
